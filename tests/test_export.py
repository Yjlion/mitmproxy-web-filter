"""Tests for GET /api/logs/export — CSV and XLSX download endpoint."""
from __future__ import annotations

import io
import sys
import time
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))

# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _reset_logstore():
    """Close any open write connection and reset module state."""
    import shared.logstore as ls
    if ls._conn is not None:
        try:
            ls._conn.close()
        except Exception:
            pass
        ls._conn = None
    ls._db_path = None
    ls._insert_count = 0


def _seed_db(tmp_path: Path) -> str:
    """Initialise a fresh logstore in tmp_path/webfilter.db (matching
    GlobalSettings.db_path) and insert a few rows. Returns the db path."""
    import shared.logstore as ls

    _reset_logstore()

    # Use "webfilter.db" — same name GlobalSettings.db_path returns.
    db = str(tmp_path / "webfilter.db")
    ls.configure(db, retention_days=0)  # retention_days=0 → keep everything

    now = int(time.time())
    for i in range(3):
        ls.log_request({
            "ts": now - (3 - i) * 60,
            "method": "GET",
            "host": f"host{i}.example.com",
            "path": f"/path{i}",
            "status": 200,
            "action": "ok",
            "component": "",
            "policy": "default",
            "client_ip": f"10.0.0.{i+1}",
        })
        ls.log_block({
            "ts": now - (3 - i) * 60,
            "domain": f"bad{i}.example.com",
            "url": f"https://bad{i}.example.com/",
            "reason": "blocklist",
            "component": "url_filter",
            "policy": "kids",
            "client_ip": f"10.0.0.{i+1}",
        })
    return db


def _make_client(tmp_path: Path):
    """Return (TestClient, main_mod, original_fn) for the given tmp_path.
    The caller must restore main_mod._load_settings = original_fn when done."""
    from fastapi.testclient import TestClient
    import management.api.main as main_mod
    from shared.models import GlobalSettings

    _seed_db(tmp_path)

    def fake_load():
        s = GlobalSettings()
        s.logs_dir = str(tmp_path)
        return s

    original = main_mod._load_settings
    main_mod._load_settings = fake_load
    client = TestClient(main_mod.app, raise_server_exceptions=True)
    return client, main_mod, original


# ---------------------------------------------------------------------------
# Tests — CSV
# ---------------------------------------------------------------------------

class TestExportCSV:
    def test_csv_requests(self, tmp_path):
        import management.api.main as main_mod
        client, main_mod, original = _make_client(tmp_path)
        try:
            resp = client.get("/api/logs/export?kind=requests&format=csv")
        finally:
            main_mod._load_settings = original

        assert resp.status_code == 200
        assert "text/csv" in resp.headers["content-type"]
        assert 'filename="requests-' in resp.headers["content-disposition"]

        lines = resp.text.strip().splitlines()
        # Header + 3 data rows
        assert len(lines) == 4
        # Header has expected columns
        header = lines[0].split(",")
        assert "host" in header
        assert "action" in header

    def test_csv_blocks(self, tmp_path):
        import management.api.main as main_mod
        client, main_mod, original = _make_client(tmp_path)
        try:
            resp = client.get("/api/logs/export?kind=blocks&format=csv")
        finally:
            main_mod._load_settings = original

        assert resp.status_code == 200
        lines = resp.text.strip().splitlines()
        assert len(lines) == 4  # header + 3 rows
        header = lines[0].split(",")
        assert "domain" in header


# ---------------------------------------------------------------------------
# Tests — XLSX
# ---------------------------------------------------------------------------

class TestExportXLSX:
    def test_xlsx_requests(self, tmp_path):
        import openpyxl
        import management.api.main as main_mod
        client, main_mod, original = _make_client(tmp_path)
        try:
            resp = client.get("/api/logs/export?kind=requests&format=xlsx")
        finally:
            main_mod._load_settings = original

        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
        assert 'filename="requests-' in resp.headers["content-disposition"]

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # header row + 3 data rows
        assert len(rows) == 4
        # First row is the header
        assert "host" in rows[0]

    def test_xlsx_blocks(self, tmp_path):
        import openpyxl
        import management.api.main as main_mod
        client, main_mod, original = _make_client(tmp_path)
        try:
            resp = client.get("/api/logs/export?kind=blocks&format=xlsx")
        finally:
            main_mod._load_settings = original

        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 4
        assert "domain" in rows[0]


# ---------------------------------------------------------------------------
# Tests — Validation
# ---------------------------------------------------------------------------

class TestExportValidation:
    def test_invalid_kind_returns_400(self, tmp_path):
        import management.api.main as main_mod
        client, main_mod, original = _make_client(tmp_path)
        try:
            resp = client.get("/api/logs/export?kind=bad_table&format=csv")
        finally:
            main_mod._load_settings = original
        assert resp.status_code == 400

    def test_invalid_format_returns_400(self, tmp_path):
        import management.api.main as main_mod
        client, main_mod, original = _make_client(tmp_path)
        try:
            resp = client.get("/api/logs/export?kind=requests&format=xml")
        finally:
            main_mod._load_settings = original
        assert resp.status_code == 400

    def test_over_cap_returns_400(self, tmp_path, monkeypatch):
        import management.api.main as main_mod
        client, main_mod, original = _make_client(tmp_path)
        # Lower the cap so our 3 rows exceed it.
        monkeypatch.setattr(main_mod, "MAX_EXPORT_ROWS", 2)
        try:
            resp = client.get("/api/logs/export?kind=requests&format=csv")
        finally:
            main_mod._load_settings = original

        assert resp.status_code == 400
        assert "limit" in resp.json()["detail"].lower()

    def test_date_range_filters_rows(self, tmp_path):
        """Only rows within [start, end] should appear in the export."""
        import management.api.main as main_mod
        import shared.logstore as ls
        from fastapi.testclient import TestClient
        from shared.models import GlobalSettings

        # Use real-time-based timestamps so they survive retention pruning.
        now = int(time.time())
        ts_early = now - 7200   # 2 hours ago
        ts_mid   = now - 3600   # 1 hour ago
        ts_late  = now - 1800   # 30 minutes ago

        # Seed with known timestamps (overwrite the default seed).
        _reset_logstore()
        db = str(tmp_path / "webfilter.db")
        ls.configure(db, retention_days=0)
        for ts in [ts_early, ts_mid, ts_late]:
            ls.log_request({"ts": ts, "host": f"h{ts}.com", "action": "ok",
                            "method": "GET", "path": "/", "status": 200,
                            "component": "", "policy": "p", "client_ip": "1.2.3.4"})

        # Export range: only the middle row should be included.
        start = ts_early + 1   # exclude early
        end   = ts_late  - 1   # exclude late

        def fake_load():
            s = GlobalSettings()
            s.logs_dir = str(tmp_path)
            return s

        original = main_mod._load_settings
        main_mod._load_settings = fake_load
        try:
            client = TestClient(main_mod.app, raise_server_exceptions=True)
            resp = client.get(f"/api/logs/export?kind=requests&format=csv&start={start}&end={end}")
        finally:
            main_mod._load_settings = original

        lines = resp.text.strip().splitlines()
        # header + 1 data row (only ts_mid is in range)
        assert len(lines) == 2
